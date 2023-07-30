import datetime
import random
import requests

from cryptography.fernet import InvalidToken
from fastapi import HTTPException, APIRouter, UploadFile, Form, File
from fastapi.responses import FileResponse
from jose.exceptions import ExpiredSignatureError
from openpyxl import Workbook
from starlette.responses import JSONResponse

from Database import Databases
from Database.Databases import NotFoundError, BadIpuDeltaError
from Models.Models import *
from Routes.Authorization import unpack_token, BadTokenError
from Utils.Hasher import HasherClass, EncryptionClass
from Utils.QRscanner import scanQR
from config import SECRET_KEY, iputoken

Database = Databases.DatabaseBaseClass()
router = APIRouter()
Hasher = HasherClass()
SQLDatabase = Databases.SQLDatabase()
Encrypter = EncryptionClass()


# --------------------FUNCTIONS--------------------

def count_sum(delta_number, tariff):
    return delta_number * tariff


async def set_verdict(user_id: int, ipu: str, new_number: str):
    info = (await SQLDatabase.get_last_values(user_id, ipu))
    info = list(reversed(info))
    counter = 0
    diff = 0
    if len(info) == 1:
        if (abs(int((datetime.datetime.now() - info[0]['date']).days) * 130 - (int(new_number) - int(info[0]['new_number'])))) / (int(new_number) - int(info[0]['new_number'])) < 0.3:
            return "Не подозрительно"
        else:
            return "Подозрительно"
    for dictionary in info:
        if counter == 0:
            first_date = dictionary['date']
            first_number = int(dictionary['new_number'])
        else:
            date_diff = int((dictionary['date'] - first_date).days)
            num_diff = int(dictionary['new_number']) - first_number
            if date_diff == 0:
                diff = 130
            else:
                diff += num_diff / date_diff
            first_date = dictionary['date']
            first_number = int(dictionary['new_number'])

        counter += 1
    if len(info) == 3:
        diff /= 2
    if int((datetime.datetime.now() - info[-1]['date']).days):
        return "Не подозрительно"
    if (abs(int((datetime.datetime.now() - info[-1]['date']).days) * diff - (int(new_number) - int(info[-1]['new_number'])))) / abs(int(new_number) - int(info[-1]['new_number'])) < 0.3:
        return "Не подозрительно"
    else:
        return "Подозрительно"


async def insert_info(func_name, username, sheet, headers):
    PAGE_SIZE = 15
    offset = 0
    used_function = getattr(SQLDatabase, func_name)
    while True:
        data = await used_function(username, offset)
        if not data:
            break
        if func_name == 'get_all_validations' or func_name == 'get_suspicious_validations':
            ids = []
            for dictionary in data:
                ids.append(dictionary["validation_id"])
            data = []
            for id in ids:
                info = await SQLDatabase.get_validation_logs(username, id)
                info['validation_id'] = id
                data.append(info)
        else:
            for dictionary in data:
                dictionary['transaction_date'] = str(dictionary['transaction_date'])
        for row in data:
            row_data = [row[col] for col in headers]
            sheet.append(row_data)
        offset += PAGE_SIZE



# ----------------------------------------------------

headers = requests.utils.default_headers()

headers.update(
    {
        'User-Agent': 'My User Agent 1.0',
    }
)


async def add_transaction(key: str, login: str, new_number: str, ipu: str):
    try:
        user_id, tariff = await SQLDatabase.get_user_id_tariff(login)
        prev_number = await SQLDatabase.get_last_number(user_id, ipu)
        if int(new_number) <= int(prev_number):
            raise BadIpuDeltaError
        trans_id = await SQLDatabase.add_transaction(user_id, prev_number, new_number, ipu,
                                                     count_sum((int(new_number) - int(prev_number)) / 1000, tariff),
                                                     await set_verdict(user_id, ipu, new_number))
        trans_id['first_value'] = False
        return JSONResponse(trans_id)
    except NotFoundError:
        trans_id = await SQLDatabase.first_value(login, ipu, new_number)
        trans_id['first_value'] = True
        return JSONResponse(trans_id)


# --------------------ROUTES-----------------------


@router.post('/trans_status', tags=['transactions'])
async def change_trans_status(key: Secret_key, trans_id: int, status: int):
    try:
        if not Hasher.verify_password(key.key, SECRET_KEY):
            raise HTTPException(status_code=403, detail='bad secret key')
        await SQLDatabase.change_status(trans_id, status)
        return HTTPException(status_code=200)
    except:
        raise HTTPException(status_code=500, detail='Server Error')


@router.post('/scan_photo', tags=['transactions'])
async def scan_photo(photo: UploadFile = File(...), key: str = Form()):
    if not Hasher.verify_password(key, SECRET_KEY):
        raise HTTPException(status_code=403, detail='bad secret key')
    content = await photo.read()
    qr_info = scanQR(content)
    if qr_info == '':
        raise HTTPException(status_code=404, detail='QR-code not found on photo')
    try:
        info = Encrypter.decrypt_qrinfo(qr_info).split(sep=';')  # договор, счетчик
    except InvalidToken:
        raise HTTPException(status_code=404, detail='QR-code wrong info')
    data = {"token": iputoken}
    files = {"upload_image": (photo.filename, content)}
    number_result = requests.post(url='https://api.waterdroplet.ru/get_number',
                                  data=data, files=files)
    if number_result.status_code == 200:
        res = number_result.json()  # цифры на счетчике
        if res == "":
            raise HTTPException(status_code=417, detail='number not found on photo')
        try:
            info = await add_transaction(key=key, login=info[0], new_number=res['number'], ipu=info[1])
            return info
        except BadIpuDeltaError:
            raise HTTPException(status_code=417, detail='New value is less than previous or its same')
    else:
        raise HTTPException(status_code=500, detail='API service Error')


@router.post('/get_transactions_logs/{page_id}', tags=['transactions'])
async def get_transactions_logs(token: Token, page_id: int):
    try:
        username, user_type = unpack_token(token.access_token)
        if user_type == "business":
            page_id -= 1
            info = await SQLDatabase.get_transactions_logs(username, page_id)
            for dictionary in info:
                dictionary['transaction_date'] = str(dictionary['transaction_date'])
            return JSONResponse(info)
        else:
            raise HTTPException(status_code=400, detail="bad user_type")
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='token expired')
    except BadTokenError:
        raise HTTPException(status_code=401, detail='bad token')


@router.post('/get_suspicious_transactions_logs/{page_id}', tags=['transactions'])
async def get_suspicious_transactions_logs(token: Token, page_id: int):
    try:
        username, user_type = unpack_token(token.access_token)
        if user_type == "business":
            page_id -= 1
            info = await SQLDatabase.get_sus_transactions_logs(username, page_id)
            for dictionary in info:
                dictionary['transaction_date'] = str(dictionary['transaction_date'])
            return JSONResponse(info)
        else:
            raise HTTPException(status_code=400, detail="bad user_type")
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='token expired')
    except BadTokenError:
        raise HTTPException(status_code=401, detail='bad token')


@router.post('/save_file', tags=['transactions'])
async def save_file(token: Token):
    try:
        username, user_type = unpack_token(token.access_token)
        if user_type == "business":
            workbook = Workbook()

            sheet = workbook.active
            sheet.title = 'Все транзакции'
            sheet.append(['№ транзакции', 'ФИО', 'Дата', 'Счетчик', 'Пред. значение', 'Новое значение', 'Сумма оплаты',
                          'Подозрительность'])
            headers = ['transaction_id', 'full_name', 'transaction_date', 'ipu', 'prev_number', 'new_number',
                       'payment_sum', 'verdict']
            await insert_info(func_name='get_transactions_logs', username=username, sheet=sheet, headers=headers)

            sheet = workbook.create_sheet('Подозрительные транзакции')
            sheet.append(['№ транзакции', 'ФИО', 'Дата', 'Счетчик', 'Пред. значение', 'Новое значение', 'Сумма оплаты',
                          'Подозрительность'])
            await insert_info(func_name='get_sus_transactions_logs', username=username, sheet=sheet, headers=headers)

            sheet = workbook.create_sheet('Проверки сотрудников')
            sheet.append(["№ проверки", "Имя сотрудника", "Дата проверки", "Значение сотрудника", "ФИО клиента",
                          "Дата показаний", "Значение клиента", "Подозрительность"])
            headers = ["validation_id", "sotrudnik_name", "sotrudnik_photo_date", "sotrudnik_number", "physic_name",
                       "physic_photo_date", "physic_number", "verdict"]
            await insert_info(func_name='get_all_validations', username=username, sheet=sheet, headers=headers)

            sheet = workbook.create_sheet('Проверки (подозрительные)')
            sheet.append(["№ проверки", "Имя сотрудника", "Дата проверки", "Значение сотрудника", "ФИО клиента",
                          "Дата показаний", "Значение клиента", "Подозрительность"])
            await insert_info(func_name='get_suspicious_validations', username=username, sheet=sheet, headers=headers)

            file_path = 'data.xlsx'
            workbook.save(file_path)
            workbook.close()
            return FileResponse(file_path, filename='data.xlsx')

        else:
            raise HTTPException(status_code=400, detail="bad user_type")
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail='token expired')
    except BadTokenError:
        raise HTTPException(status_code=401, detail='bad token')
